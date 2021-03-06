from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.urls import reverse
from django.core.paginator import Paginator
from . import models
from . import utils
from django.conf import settings as my_settings
import openpyxl
from openpyxl.utils import get_column_letter
import datetime


# тут только "логика" - функции для обработки и возврат данных

# class Todo:
#     def __init__(self, description, name="name1"):
#         self.description = description
#         self.name = name
#         self.value = 0
#
#     def counter(self, external_value):
#         self.value += external_value
#
#     @staticmethod
#     def count(value, extra):
#         return (value + extra) * 1000


#
# obj = Todo("sdomethi", "name2")
#
# obj.description
# obj.name

# Todo.counter()
# Todo.count()

def external_excel():
    def create_new_excel():
        workbook = openpyxl.Workbook()
        # grab the active worksheet
        worksheet = workbook.active
        # Data can be assigned directly to cells
        worksheet['A1'] = 42
        # Rows can also be appended
        worksheet.append([1, 2, 3])
        # Python types will automatically be converted
        worksheet['A2'] = datetime.datetime.now()
        # Save the file
        workbook.save("./static/temp/sample.xlsx")

    def load_excel():
        path = "./static/temp/sample.xlsx"
        workbook = openpyxl.load_workbook(path)
        worksheet = workbook.active
        max_num_rows = worksheet.max_row

        global_list = []
        for num in range(1, max_num_rows):
            local_list = []
            for char in "ABC":
                local_list.append(worksheet[f'{char}{num}'].value)
            global_list.append(local_list)

    def update_excel():
        path = "./static/temp/sample.xlsx"
        workbook = openpyxl.load_workbook(path)
        worksheet = workbook.active
        max_num_rows = worksheet.max_row

        global_list = []
        for num in range(1, max_num_rows):
            local_list = []
            for char in "ABC":
                local_list.append(worksheet[f'{char}{num}'].value)
            global_list.append(local_list)

        workbook = openpyxl.Workbook()
        # grab the active worksheet
        worksheet = workbook.active

        index_i = 0
        for i in global_list:
            index_i += 1
            index_j = 0
            for j in i:
                index_j += 1
                worksheet[f'{get_column_letter(index_i)}{index_j}'] = str(j) + " Bogdan"

        # Save the file
        workbook.save("./static/temp/sample.xlsx")

    def beatiful_new_excel():  # TODO нужно реализовать !!!!!!!!!!!
        pass

    update_excel()


def index(request):
    if my_settings.DEBUG:
        context = {"name": "Ally", "age": 25}
    else:
        context = {"name": "Ally", "age": 22}

    return HttpResponse()


def home(request):
    external_excel()
    return render(request, 'app_teacher/pages/home.html')
    # return render(request, 'index2.html')


def about(request):
    return render(request, 'app_teacher/pages/about.html')


def origin_home(request):
    return render(request, 'app_teacher/pages/origin_home.html')


def todo_detail(request, todo_id):
    obj = models.Task.objects.get(id=todo_id)
    context = {
        "todo": obj
    }
    return render(request, 'app_teacher/pages/DetailTodo.html', context)


def todo_list(request):
    objects = models.Task.objects.all()
    count_object_on_one_page = 2
    current_page_from_request_parametr = request.GET.get('page')
    page_obj = utils.CustomPaginator.get_page(
        objs=objects,
        limit=count_object_on_one_page,
        current_page=current_page_from_request_parametr
    )
    context = {"list": None, "page": page_obj, "iterator": range(0, 20),
               "value": [11274.25234533463, 1474.25234463, 174.2523453463],
               "values": [3274.00, 14232374.00, 23441234.00], "new": "Bananas"}
    return render(request, 'app_teacher/pages/todo_list.html', context)


def todo_create(request):
    if request.method == "POST":
        title1 = request.POST.get("title", "заголовок по умолчанию")
        description1 = request.POST.get("description", "описание по умолчанию")
        obj = models.Task.objects.create(
            title=title1,
            description=description1
        )
        obj.save()
    context = {}
    return render(request, 'app_teacher/pages/CreateTodo.html', context)


def todo_delete(request, todo_id):
    obj = models.Task.objects.get(id=todo_id)
    obj.delete()
    return redirect(reverse('todo_list', args=()))


def todo_update_status(request, todo_id):
    obj = models.Task.objects.get(id=todo_id)
    # obj.is_completed = not obj.is_completed
    if obj.is_completed:
        obj.is_completed = False
    else:
        obj.is_completed = True
    obj.save()
    return redirect(reverse('todo_list', args=()))


def todo_change_data(request, todo_id):
    obj = models.Task.objects.get(id=todo_id)

    if request.method == "POST":
        title1 = request.POST.get("title", "заголовок по умолчанию")
        description1 = request.POST.get("description", "описание по умолчанию")
        if obj.title != title1:
            obj.title = title1
        if obj.description != description1:
            obj.description = description1
        obj.save()
    context = {
        "todo": obj
    }
    return render(request, 'app_teacher/pages/ChangeTodo.html', context)


def admin_page(request):
    if request.method == "POST":
        excel = request.FILES.get("excel", None)
        print(excel)
        workbook = openpyxl.load_workbook(excel)
        sheet = workbook.active
        # local_value = sheet['B2'].value

        global_list = []
        for num in range(1, 20 + 1):
            local_list = []
            for char in "ABC":
                local_list.append(sheet[f'{char}{num}'].value)
            global_list.append(local_list)
        print("global_list: ", global_list)

        for i in global_list:
            obj = models.Task.objects.create(
                title=i[1],
                description=i[2]
            )
    context = {
    }
    return render(request, 'app_teacher/pages/AdminPage.html', context)
